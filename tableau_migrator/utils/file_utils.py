import csv

from collections import defaultdict
from openpyxl import load_workbook
from pathlib import Path


def load_workbook_luids(path: Path | str, sheet: str) -> list[str]:
    wb = load_workbook(path, read_only=True)
    ws = wb[sheet]
    header = [c.value for c in next(ws.iter_rows(max_row=1))]
    luid_idx = header.index("luid")
    return [
        row[luid_idx]
        for row in ws.iter_rows(min_row=2, values_only=True)
        if row[luid_idx]
    ]


def load_credentials(path: Path | str) -> dict[str, list[dict[str, str]]]:
    creds = defaultdict(list)
    with open(path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            creds[row["dbclass"].lower()].append(row)
    return creds
